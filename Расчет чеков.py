import json
import openpyxl as oxl
import os
from glob import glob
from pathlib import Path
import datetime
from openpyxl.styles import Side,Border,PatternFill
import sys
#print(sys.path[0])
excel_path=os.path.dirname(os.path.realpath(__file__))
#print(excel_path)
first_folder=Path('C:/Users/Wladimir/Downloads')
month_colour=['FF0000','FF7F00','FFFF00','7FFF00','00FF00','00FF7F','00FFFF','007FFF','0000FF','7F00FF','FF00FF','FF007F',]
category_colour={'для дома':"D8BFD8",'молочные продукты':"ADD8E6",'овощи и фрукты':"D8E4BC",'чай и сладкое':"D2691E",
                 'снеки':"FFA500",'бакалея':"CD5C5C",'мясо и птица':"FFA07A",'здоровье':"8FBC8F",'кафе':"CD853F",
                 'канцтовары':"708090",'обед':"8B4513",'косметика':"E6E6FA",'упаковка':"C0C0C0",'напитки':"AFEEEE",
                 'гастрономия':"FF7F50",'хлеб':"F5DEB3"}

def getfiles(dirname):
    return glob(os.path.join(dirname, '*.json'))

paths=getfiles(first_folder)

cheks_amount=len(paths)
#print(cheks_amount)
def write_check(check_path):
    #print(check_path)
    with open(check_path, "r", encoding='utf-8') as read_file:
        chek = json.load(read_file)
    #print(chek)
    chek_date = chek['date'][0:8] + '20' + chek['date'][8:]
    # print(chek_date)
    date_datetime = datetime.datetime.strptime(chek_date, "%d.%m.%Y %H:%M")
    # print(date_datetime)
    chek_month_year = chek_date[3:8]
    # print(chek_date)
    # print(chek_month_year)
    print(chek_month_year[0:2])
    try:
        chek_sheet = wb_cheki[chek_month_year]
    except KeyError:
        chek_sheet = wb_cheki.create_sheet(chek_month_year)
        chek_sheet.cell(row=1, column=1, value='Короткое название')
        chek_sheet.cell(row=1, column=2, value='Название')
        chek_sheet.cell(row=1, column=3, value='Цена')
        chek_sheet.cell(row=1, column=4, value='Количество')
        chek_sheet.cell(row=1, column=5, value='Стоимость')
        chek_sheet.cell(row=1, column=6, value='Категория')
        chek_sheet.cell(row=1, column=7, value='Магазин')
        chek_sheet.cell(row=1, column=8, value='Адрес')
        chek_sheet.cell(row=1, column=9, value='Время')

    #check_colour = str(hex(int(chek_month_year[0:2]) * 2))[2:4] + str(hex(int(chek_month_year[0:2]) * 20))[2:4] + str(
        #hex(int(chek_month_year[0:2]) * 20))[2:4])
    check_colour=month_colour[-int(chek_month_year[0:2])+9]
    chek_sheet.sheet_properties.tabColor = check_colour

    lastRow = chek_sheet.max_row
    #print('lastrow=' + str(lastRow))
    takiezhe = 0
    rows_takiezhe = []
    for rownumber in range(1, lastRow + 1):
        cell_i = chek_sheet.cell(row=rownumber, column=9).value
        #print('cell_i=' + str(cell_i))
        if cell_i == chek['date']:
            takiezhe = takiezhe + 1
            rows_takiezhe = rows_takiezhe + [rownumber]
    print('Магазин в чеке = ' + str(chek['shopName']))
    print('количетсво таких же строк= ' + str(takiezhe))
    # print(chek_sheet.cell(row=4, column=3).value)
    print('номера таких же строк =' + str(rows_takiezhe))
    # print(chek['shopName'])

    if takiezhe == 0:
        write_in_row = lastRow + 1
    else:
        write_in_row = rows_takiezhe[0]
    print('Пишу в строке ='+ str(write_in_row))

    def write_data_chek(row_chek):
        products_amount = len(chek['products'])
        print('Число продуктов в чеке ='+ str(products_amount))
        print('_____________________________________')

        for i in range(0, products_amount):
            chek_sheet.cell(row=row_chek + i, column=1).value = chek['products'][i]['name'].split(' ')[0]
            chek_sheet.cell(row=row_chek + i, column=2).value = chek['products'][i]['name']
            chek_sheet.cell(row=row_chek + i, column=3).value = chek['products'][i]['price']
            chek_sheet.cell(row=row_chek + i, column=4).value = chek['products'][i]['quantity']
            chek_sheet.cell(row=row_chek + i, column=5).value = chek['products'][i]['sum']
            chek_sheet.cell(row=row_chek + i, column=6).value = chek['products'][i]['category']
            chek_sheet.cell(row=row_chek + i, column=7).value = chek['shopName']
            chek_sheet.cell(row=row_chek + i, column=8).value = chek['shopAddress']
            chek_sheet.cell(row=row_chek + i, column=9).value = chek['date']
    write_data_chek(write_in_row)
os.chdir(excel_path)
wb_cheki = oxl.load_workbook('Чеки.xlsx')
for n in range(0,cheks_amount):
    write_check(paths[n])
for sheet in wb_cheki:
    lastRow=sheet.max_row
    for rownumber in range(1, lastRow + 1):
        row_category = str(sheet.cell(row=rownumber, column=6).value).lower()
        if row_category == "категория":
            for m in range(1, 9 + 1):
                #print("котегория")
                sheet.cell(row=rownumber, column=m).fill = PatternFill(fgColor="FFFFFF",fill_type="solid")
                bd = Side(style='thick', color="000000")
                sheet.cell(row=rownumber, column=m).border = Border(left=bd, top=bd, right=bd, bottom=bd)
        if row_category in category_colour:
            for m in range(1, 9 + 1):
                sheet.cell(row=rownumber, column=m).fill = PatternFill(fgColor=category_colour[row_category],fill_type="solid")
                bd = Side(style='thin', color="C0C0C0")
                sheet.cell(row=rownumber, column=m).border = Border(left=bd, top=bd, right=bd, bottom=bd)

        else:
            for m in range(1, 9 + 1):
                sheet.cell(row=rownumber, column=m).fill = PatternFill(fgColor="FFFFFF", fill_type="solid")
                bd = Side(style='thin', color="C0C0C0")
                sheet.cell(row=rownumber, column=m).border = Border(left=bd, top=bd, right=bd, bottom=bd)


wb_cheki.save('Чеки.xlsx')
print("Готово")
input()
